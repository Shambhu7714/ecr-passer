import pandas as pd
import openpyxl
import os
import re
import numpy as np
from core.country_config import CountryConfigLoader

_country_loader = CountryConfigLoader()

def preprocess_excel(file_path, sheet_name=None, mapping_file=None):
    """
    Reads raw Excel, handles merged cells, and prepares a clean grid.
    Removes header rows, footers, and extracts only the data table.
    Flattens multi-level headers properly.
    """
    country_cfg = _country_loader.detect_and_load(file_path, mapping_file)
    header_keywords = country_cfg.get("header_detection_keywords", [])
    footer_keywords_cfg = country_cfg.get("footer_keywords", [])
    
    effective_sheet = sheet_name if sheet_name is not None else 0
    df = pd.read_excel(file_path, sheet_name=effective_sheet, header=None)
    
    header_row_idx = _find_header_by_keywords(df, header_keywords)
    
    # If not found, look for time-axis headers (pivot detection)
    if header_row_idx is None:
        time_kws = country_cfg.get("time_keywords", [])
        for idx, row in df.iterrows():
            row_vals = [str(x).lower() for x in row.values if pd.notna(x)]
            hits = sum(1 for v in row_vals if any(tk.lower() in v for tk in time_kws) or (v.isdigit() and 1990 < int(v) < 2050))
            if hits >= 3:
                header_row_idx = idx
                break
                
    if header_row_idx is None:
        header_row_idx = _find_data_table_start(df)
        if header_row_idx is None:
            header_row_idx = 0

    second_header_idx = header_row_idx + 1
    has_second_header = False

    # Check if a second header exists (for multi-level headers)
    # Detect if second row is also part of header (e.g. Row 7 is Year, Row 8 is Month)
    if second_header_idx < len(df):
        second_row = df.iloc[second_header_idx]
        second_vals = [str(v).strip() for v in second_row.values if pd.notna(v) and str(v).strip() not in ('', 'nan')]
        
        # If second row has many months or years, it's a second header
        time_kws = country_cfg.get("time_keywords", [])
        second_hits = sum(1 for v in second_vals if any(tk.lower() in v.lower() for tk in time_kws) or (v.isdigit() and 1990 < int(v) < 2050))
        
        if second_hits >= 3 or any(kw in ' '.join(second_vals).lower() for kw in ['municipio', 'total', 'área']):
            has_second_header = True

    if has_second_header:
        df_clean = _read_excel_with_merged_headers(file_path, effective_sheet, header_row_idx, second_header_idx)
    else:
        df_clean = pd.read_excel(file_path, sheet_name=effective_sheet, header=header_row_idx)

    df_clean = df_clean.dropna(how='all').dropna(how='all', axis=1)
    
    # Simple footer cleanup
    if len(df_clean) > 3:
        for idx_footer, row_footer in df_clean.iterrows():
            search_area = " ".join(str(val) for val in row_footer.iloc[:3].values if pd.notna(val)).lower()
            if any(kw.lower() in search_area for kw in footer_keywords_cfg):
                df_clean = df_clean.iloc[:idx_footer]
                break

    # Propagate labels for merged/sparse cells (crucial for Year/Category columns)
    # Aggressively clean up empty-looking strings so ffill works
    df_clean = df_clean.map(lambda x: np.nan if str(x).strip().lower() in ['', 'nan', 'none', 'null'] else x)
    df_clean = df_clean.ffill(axis=0)

    return df_clean

def _read_excel_with_merged_headers(file_path, sheet_name, header_row_idx, second_header_idx):
    # Use data_only=True to get values instead of formulas
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    # Resolve merged cells in memory
    merged_ranges = list(ws.merged_cells.ranges)
    for m_range in merged_ranges:
        min_col, min_row, max_col, max_row = m_range.min_col, m_range.min_row, m_range.max_col, m_range.max_row
        top_left_value = ws.cell(row=min_row, column=min_col).value
        # If the merged cell is in or near our header area, propagate it
        if min_row <= second_header_idx + 2:
            ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_left_value

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    top_row = list(all_rows[header_row_idx])
    sub_row = list(all_rows[second_header_idx])

    max_col_idx = 0
    for i, v in enumerate(top_row):
        if v is not None: max_col_idx = i
    for i, v in enumerate(sub_row):
        if v is not None: max_col_idx = max(max_col_idx, i)
    n_cols = max_col_idx + 1

    top_row = top_row[:n_cols]
    sub_row = sub_row[:n_cols]

    # Convert to Series for ffill
    # Ensure empty strings are treated as NaN for ffilling
    top_s = pd.Series(top_row).replace('', np.nan).replace('None', np.nan).ffill()
    sub_s = pd.Series(sub_row).replace('', np.nan).replace('None', np.nan)

    def _clean(s):
        return ' '.join(str(s).split()) if pd.notna(s) and str(s).strip() != 'None' else ''

    columns = []
    for i in range(n_cols):
        t = _clean(top_s.iloc[i])
        s = _clean(sub_s.iloc[i])
        if t and s and t.lower() != s.lower():
            columns.append(f"{t} | {s}")
        elif t:
            columns.append(t)
        elif s:
            columns.append(s)
        else:
            columns.append(f"Col_{i}")

    # Unique columns
    seen = {}
    unique_cols = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            unique_cols.append(f"{col}.{seen[col]}")
        else:
            seen[col] = 0
            unique_cols.append(col)

    data_rows = all_rows[second_header_idx + 1:]
    df_out = pd.DataFrame([list(r)[:n_cols] for r in data_rows], columns=unique_cols)
    return df_out

def _find_header_by_keywords(df, keywords):
    best_row_idx = None
    max_hits = 0
    
    # Check first 50 rows for header
    for idx, row in df.head(50).iterrows():
        row_vals = [str(x).strip() for x in row.values if pd.notna(x)]
        if not row_vals: continue
        
        row_str = ' '.join(row_vals).lower()
        
        # If row is very long (title), it's likely not the header row
        if len(row_str) > 200: continue
        
        hits = 0
        for kw in keywords:
            pattern = r'\b' + re.escape(kw.lower()) + r'\b'
            if re.search(pattern, row_str):
                hits += 1
        
        # We want the row with the MOST keyword hits
        if hits > max_hits:
            max_hits = hits
            best_row_idx = idx
            
    # Require at least 2 hits for structural headers if we found something
    # or return the best if we have high confidence
    if max_hits >= 1:
        return best_row_idx
        
    return None

def _find_data_table_start(df):
    for idx, row in df.iterrows():
        non_null = [v for v in row.values if pd.notna(v)]
        if len(non_null) >= 3:
            numeric = sum(1 for v in non_null if _is_numeric(v))
            if numeric / len(non_null) > 0.4:
                return idx - 1 if idx > 0 else 0
    return None

def _is_numeric(val):
    if isinstance(val, (int, float)):
        return not pd.isna(val)
    try:
        s = str(val).strip().replace(',', '.')
        if s == '-': return True # Sentinel
        float(s)
        return True
    except:
        return False

def save_intermediate_grid(df, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_excel(path, index=False)
