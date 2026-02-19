"""
Smart Sampler - Efficiently samples large files without loading everything.
Part of the Hybrid Agentic Parser implementation.
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Tuple
import hashlib


class SmartSampler:
    """
    Efficiently samples data from large files without loading everything.
    Extracts corners, middle, and bottom sections to understand structure.
    """
    
    def __init__(self, sample_size: int = 5):
        """
        Initialize the smart sampler.
        
        Args:
            sample_size: Number of rows/columns to sample from each region
        """
        self.sample_size = sample_size
        
    def sample_excel(self, file_path: str, sheet_name: str = None) -> Dict[str, pd.DataFrame]:
        """
        Sample an Excel sheet intelligently without loading the entire file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to sample (if None, uses first sheet)
            
        Returns:
            Dict with 'top_left', 'middle', 'bottom' samples as DataFrames
        """
        # Open workbook in read-only mode for efficiency
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Get the sheet
        if sheet_name is None:
            ws = wb.active
        else:
            ws = wb[sheet_name]
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"  File dimensions: {max_row} rows   {max_col} columns")
        
        samples = {}
        
        # Top-left corner (usually contains headers + first data rows)
        print(f"  Sampling top-left corner ({self.sample_size + 5} rows   {min(max_col, 20)} cols)...")
        samples['top_left'] = self._extract_region(
            ws, 
            1, 1, 
            min(self.sample_size + 5, max_row), 
            min(max_col, 20)
        )
        
        # Middle section (representative data)
        if max_row > 20:
            mid_row = max_row // 2
            print(f"  Sampling middle section (rows {mid_row-2} to {mid_row+2})...")
            samples['middle'] = self._extract_region(
                ws, 
                max(mid_row - 2, 1), 
                1,
                min(mid_row + 2, max_row),
                min(max_col, 20)
            )
        else:
            samples['middle'] = samples['top_left'].copy()
        
        # Bottom section (check for summaries/totals)
        if max_row > 10:
            print(f"  Sampling bottom section (last {min(3, max_row)} rows)...")
            samples['bottom'] = self._extract_region(
                ws,
                max(max_row - 3, 1),
                1,
                max_row,
                min(max_col, 20)
            )
        else:
            samples['bottom'] = samples['top_left'].copy()
        
        wb.close()
        
        print(f"[OK] Sampling complete - loaded {sum(len(df) for df in samples.values())} total rows (vs {max_row} in file)")
        
        return samples
    
    def _extract_region(self, worksheet, start_row: int, start_col: int, 
                       end_row: int, end_col: int) -> pd.DataFrame:
        """Extract a region from worksheet and convert to DataFrame."""
        data = []
        for row in worksheet.iter_rows(min_row=start_row, max_row=end_row,
                                      min_col=start_col, max_col=end_col):
            data.append([cell.value for cell in row])
        
        if not data:
            return pd.DataFrame()
        
        # First row as headers
        df = pd.DataFrame(data[1:], columns=data[0])
        return df
    
    def get_structure_signature(self, samples: Dict[str, pd.DataFrame]) -> str:
        """
        Create a unique signature for the file structure.
        Used to check if we've seen this pattern before (fast-track routing).
        
        Args:
            samples: Dictionary of sampled DataFrames
            
        Returns:
            Unique signature string
        """
        top = samples['top_left']
        
        if top.empty:
            return "empty_file"
        
        signature_parts = []
        
        # 1. Column count
        signature_parts.append(f"cols:{len(top.columns)}")
        
        # 2. Header pattern (first few column names)
        headers = [str(h) for h in top.columns[:5]]
        header_hash = hashlib.md5('|'.join(headers).encode()).hexdigest()[:8]
        signature_parts.append(f"head:{header_hash}")
        
        # 3. Data type pattern in first data rows
        if len(top) > 0:
            type_pattern = ''
            for i in range(min(10, len(top.columns))):
                col_data = top.iloc[:, i]
                if pd.api.types.is_numeric_dtype(col_data):
                    type_pattern += 'N'
                elif pd.api.types.is_datetime64_any_dtype(col_data):
                    type_pattern += 'D'
                else:
                    type_pattern += 'S'
            signature_parts.append(f"types:{type_pattern}")
        
        # 4. Row count range
        row_range = "small" if len(top) < 20 else "medium" if len(top) < 100 else "large"
        signature_parts.append(f"size:{row_range}")
        
        signature = '|'.join(signature_parts)
        print(f"  Generated signature: {signature}")
        
        return signature
    
    def is_known_pattern(self, signature: str, known_patterns: Dict) -> Tuple[bool, Dict]:
        """
        Check if this signature matches a known pattern.
        
        Args:
            signature: The file signature to check
            known_patterns: Dictionary of known patterns
            
        Returns:
            (is_known, pattern_info) tuple
        """
        if signature in known_patterns:
            print(f"[OK] Matched known pattern: {signature}")
            return True, known_patterns[signature]
        
        print(f"[WARN]  Unknown pattern - will use AI analysis")
        return False, {}
